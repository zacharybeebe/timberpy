import os
from copy import deepcopy
from csv import reader, excel
from openpyxl import load_workbook
from xlrd import open_workbook
from timberpy.inventory.inventory_config import *
from timberpy.inventory.inventory_exceptions import InventoryImportError


class Inventory:
    _COL_SEPARATORS = [' ', '_', '-', '.', '']
    _TREE_ATTRS_IS_ERROR = {
        0: lambda value: Inventory._error_float(value, non_negative=False),
        1: lambda value: Inventory._error_int(value),
        2: lambda value: Inventory._error_string(value),
        3: lambda value: Inventory._error_float(value),
        4: lambda value: Inventory._error_float(value, required=False),
    }
    _LOG_ATTRS_IS_ERROR = {
        0: lambda value: Inventory._error_int(value),
        1: lambda value: Inventory._error_int(value),
        2: lambda value: Inventory._error_string(value, required=False),
        3: lambda value: Inventory._error_int(value, required=False)
    }

    def __init__(self, filename):
        self.readers = {
            '.xlsx': self.read_xlsx,
            '.xls': self.read_xls,
            '.csv': self.read_csv
        }

        if not os.path.isfile(filename):
            raise InventoryImportError('not found', filename)

        elif not any([filename.endswith(i) for i in self.readers.keys()]):
            raise InventoryImportError('not ext', filename)

        self.required = {}

        self.filename = filename
        self.ext = f'.{self.filename.split(".")[-1]}'
        self.data_read = self.readers[self.ext]()

        # for stand in self.data_read:
        #     print(f'{stand=}')
        #     print('hdr=', self.data_read[stand]['hdr'])
        #     for plot in self.data_read[stand]['plots']:
        #         print(f'\t{plot=}')
        #         print(f'\tplot_factor={self.data_read[stand]["plots"][plot]["plot_factor"]}')
        #         for row in self.data_read[stand]["plots"][plot]['trees']:
        #             print(f'\t\t{row}')
        #     print()

    def read_xlsx(self):
        # Using openpyxl library
        wb = load_workbook(self.filename, data_only=True)
        full_cruise, found_sheet = self._find_sheet(wb)
        ws = wb[found_sheet]
        return self._gather_sheet_data(ws, full_cruise)

    def read_xls(self):
        # Using xlrd library
        wb = open_workbook(self.filename)
        full_cruise, found_sheet = self._find_sheet(wb)

        ws = wb.sheet_by_name(found_sheet)
        return self._gather_sheet_data(ws, full_cruise)

    def read_csv(self):
        with open(self.filename, 'r') as csv_file:
            csv_read = reader(csv_file, dialect=excel)
            full_cruise, _ = self._find_sheet(csv_read)
            return self._gather_sheet_data(csv_read, full_cruise)

    def _find_sheet(self, workbook):
        # For .xlsx and .xls workbook object is passed. For .csv csv_read is passed

        full_cruise = True
        if self.ext in ['.xlsx', '.xls']:
            # For Excel Files
            # iter_func_to_eval and sheet_func_to_eval are the callables specific to either the openpyxl library or xlrd

            iter_func_to_eval = 'workbook.sheetnames' if self.ext == '.xlsx'            else 'workbook.sheet_names()'
            sheet_func_to_eval = 'workbook.get_sheet_by_name' if self.ext == '.xlsx'    else 'workbook.sheet_by_name'
            idx_offset = 1 if self.ext == '.xlsx'                                       else 0
            full_cruise = True
            found_sheet = None
            for sheet in eval(iter_func_to_eval):
                self.required = {col_name: REQUIRED_COLS[col_name] for col_name in REQUIRED_COLS}
                ws = eval(sheet_func_to_eval)(sheet)
                has_required, headers = self._check_sheet_for_required(ws, idx_offset)
                if not has_required:
                    # Continue looking in other sheets
                    continue
                else:
                    # Check if there are any log columns in the sheet, else default to quick cruise (automated log scali
                    other_cols = self._get_log_cols(headers, idx_offset)
                    if not other_cols:
                        full_cruise = False
                        other_cols = self._get_quick_cols(headers, idx_offset)
                    self.required.update(**other_cols)
                    found_sheet = sheet
                    return full_cruise, found_sheet
        else:
            # For CSV Files
            has_required, headers = self._check_sheet_for_required(workbook)
            if has_required:
                other_cols = self._get_log_cols(headers)
                if not other_cols:
                    full_cruise = False
                    other_cols = self._get_quick_cols(headers)
                self.required.update(**other_cols)
                return full_cruise, None

        # Either logical expression will return if required columns are found otherwise raise InventoryImportError
        raise InventoryImportError('not req', self.filename)

    def _check_sheet_for_required(self, sheet, idx_offset=0):
        self.required = {col_name: REQUIRED_COLS[col_name] for col_name in REQUIRED_COLS}
        if self.ext == '.xlsx':
            headers = [i.upper() if i is not None else i for i in next(sheet.values)]
        elif self.ext == '.xls':
            headers = [i.value.upper() if i is not None else i for i in sheet.row(0)]
        else:
            headers = [i.upper() for i in next(sheet)]
        has_required = self._check_required_cols(headers, idx_offset=idx_offset)
        return has_required, headers

    def _gather_sheet_data(self, sheet, full_cruise):
        if self.ext == '.xlsx':
            iter_func_to_eval = 'range(2, sheet.max_row + 1)'
            cell_val_to_eval = 'sheet.cell(row, idx).value'
        elif self.ext == '.xls':
            iter_func_to_eval = 'range(1, sheet.nrows)'
            cell_val_to_eval = 'sheet.row(row)[idx].value'
        else:
            iter_func_to_eval = 'sheet'
            cell_val_to_eval = 'row[idx]'

        master = {}
        stand, plot_factor, plot = None, None, None

        for row in eval(iter_func_to_eval):
            temp, all_logs, logs = [], [], []

            for col_name in self.required:
                idx = self.required[col_name]['idx']

                if col_name == 'Stand':
                    stand = eval(cell_val_to_eval)
                    if stand not in master:
                        master[stand] = {
                            'hdr': None,
                            'plots': {}
                        }
                elif col_name == 'Plot Factor':
                    plot_factor = eval(cell_val_to_eval)
                    temp.append(plot_factor)

                elif col_name == 'Plot':
                    plot = eval(cell_val_to_eval)
                    if plot not in master[stand]['plots']:
                        master[stand]['plots'][plot] = {
                            'plot_factor': plot_factor,
                            'trees': []
                        }
                else:
                    append = logs if col_name.startswith('Log') else temp
                    to_append = self.required[col_name]['default']

                    if idx is not None:
                        cell = eval(cell_val_to_eval)
                        if cell not in ['', ' ', None]:
                            to_append = cell
                    append.append(to_append)

                    if col_name.endswith('Defect') and any(logs):
                        all_logs.append(logs)
                        logs = []

            if all_logs:
                log_error, all_logs = self._check_logs(all_logs)
                if log_error:
                    raise InventoryImportError('not logs', self.filename)
                temp.append(all_logs)
                temp.append(False)

            elif not all_logs and full_cruise:
                for col_name in QUICK_COLS:
                    temp.append(QUICK_COLS[col_name]['default'])
                temp.append(True)

            else:
                temp.append(True)

            master[stand]['plots'][plot]['trees'].append(temp)

        for delete in ['', ' ', None]:
            if delete in master:
                del master[delete]

        self._check_datatypes(master)
        return master

    def _check_datatypes(self, import_dict):
        """
        :param import_dict:         The final dictionary of data gathered from the sheet
        :return:                    None

        This method will check the imported data for correct datatypes, if a datatype is incorrect for example if DBH = 'abc' it will
        raise an InventoryImportError

        If all datatypes check out it will also calculate the average HDR (height-to-diameter ratio) of the stand to fill in missing tree heights
        """
        for stand in import_dict:
            heights = []
            dbhs = []
            for plot in import_dict[stand]['plots']:
                if self._error_int(plot):
                    raise InventoryImportError('not data', self.filename)

                if self._error_float(import_dict[stand]['plots'][plot]['plot_factor'], non_negative=False):
                    raise InventoryImportError('not data', self.filename)

                for tree in import_dict[stand]['plots'][plot]['trees']:
                    for idx, err_func in self._TREE_ATTRS_IS_ERROR.items():
                        if err_func(tree[idx]):
                            raise InventoryImportError('not data', self.filename)
                        if idx == 3:
                            if tree[4] not in ['', ' ', None]:
                                heights.append(float(tree[4]))
                                dbhs.append(float(tree[idx]))
                    if not tree[-1]:
                        for log in tree[-2]:
                            for idx, err_func in self._LOG_ATTRS_IS_ERROR.items():
                                if err_func(log[idx]):
                                    raise InventoryImportError('not data', self.filename)
            if not heights:
                raise InventoryImportError('not hgt', self.filename, stand=stand)
            else:
                hdrs = [height / (dbh / 12) for height, dbh in zip(heights, dbhs)]
                import_dict[stand]['hdr'] = sum(hdrs) / len(hdrs)

    def _check_required_cols(self, headers, idx_offset=0):
        """
        :param headers:         The Column names from the inventory sheet (the first row of the sheet)
        :param idx_offset:      The index offset between the sheet and the list, either 0 (default) or 1
        :return:                A boolean of has_required

        Will look to see if the required columns are in the sheet, if they are not, Inventory will raise InventoryImportError
        """
        has_required = True
        for col_name in self.required:
            pos_cols = self._get_possible_col_names(self.required[col_name]['iters'])
            full_cols = headers + pos_cols
            filtered = list(filter(lambda x: True if full_cols.count(x) == 2 and x in pos_cols else False, full_cols))
            if filtered:
                self.required[col_name]['idx'] = headers.index(filtered[0]) + idx_offset
            else:
                has_required = False
                break
        return has_required

    def _get_log_cols(self, headers, idx_offset=0):
        """
        :param headers:         The Column names from the inventory sheet (the first row of the sheet)
        :param idx_offset:      The index offset between the sheet and the list, either 0 (default) or 1
        :return:                A log cols dictionary containing all the log numbers within the sheet and the index of those
                                log data columns

        Will look for up to 20 logs, but will break out of the loop if a given log number can't be found (if all(blanks))
        Deep copies the base LOG COLS dict and formats the main keys with the log number i.e. Log 1 Length, Log 3 Grade, ...
        Also updates the column index at which those log data are found
        """
        log_cols = {}
        for log_num in range(1, 21):
            temp_cols = {}
            blanks = [False for _ in range(len(LOG_COLS))]
            temp_cols.update({col_name.format(log_num): deepcopy(LOG_COLS[col_name]) for col_name in LOG_COLS})
            for i, col_name in enumerate(temp_cols):
                pos_cols = self._get_possible_col_names(temp_cols[col_name]['iters'], fill=log_num)
                full_cols = headers + pos_cols
                filtered = list(filter(lambda x: True if full_cols.count(x) == 2 and x in pos_cols else False, full_cols))
                if filtered:
                    temp_cols[col_name]['idx'] = headers.index(filtered[0]) + idx_offset
                else:
                    blanks[i] = True
            if all(blanks):
                break
            else:
                log_cols.update(**temp_cols)
        return log_cols

    @classmethod
    def _get_quick_cols(cls, headers: list, idx_offset: int = 0):
        """
        :param headers:         The Column names from the inventory sheet (the first row of the sheet)
        :param idx_offset:      The index offset between the sheet and the list, either 0 (default) or 1
        :return:                An updated quick cols dictionary with 'idx' sub-key updated, idx is the column index within the sheet

        Checks if the any of the quick col columns are within the sheet and updates that column index in the quick cols dictionary
        """
        quick_cols = {col_name: QUICK_COLS[col_name] for col_name in QUICK_COLS}
        for col_name in quick_cols:
            pos_cols = cls._get_possible_col_names(quick_cols[col_name]['iters'])
            full_cols = headers + pos_cols
            filtered = list(filter(lambda x: True if full_cols.count(x) == 2 and x in pos_cols else False, full_cols))
            if filtered:
                quick_cols[col_name]['idx'] = headers.index(filtered[0]) + idx_offset
        return quick_cols

    @classmethod
    def _get_possible_col_names(cls, iterators, fill=None):
        """
        :param iterators:       A list of base column names that will be concatenated with a seperator to form a possible column name
                                For example [['One', 'First'], ['Two', 'Last']] will turn into 'One_Two', 'One.Last', 'First Two', 'FirstLast', ...
        :param fill:            Use fill to fill in possible log number vals (Log Length 1, Seg_3_Grade, ...)
        :return:                A list-set of possible column names to check against the inventory sheet

        This function returns a list-set of potential column names that correspond to a particular datum, different users may use slightly different
        terminology in their inventory sheets to describe the data, so this will try to produce many logical combinations from the base names
        in the iterators arg, and see if any of those match up with any of the column names from the sheet.
        """

        possible_col_names = []
        len_iter = len(iterators)
        if fill is None:
            if len_iter == 1:
                possible_col_names += iterators[0]
            elif len_iter == 2:
                for sep in cls._COL_SEPARATORS:
                    for i in iterators[0]:
                        for j in iterators[1]:
                            possible_col_names.append(f'{i}{sep}{j}')
            else:
                for sep in cls._COL_SEPARATORS:
                    for i in iterators[0]:
                        for j in iterators[1]:
                            for k in iterators[2]:
                                possible_col_names.append(f'{i}{sep}{j}{sep}{k}')
        else:
            if len_iter == 2:
                for sep in cls._COL_SEPARATORS:
                    for i in iterators[0]:
                        for j in iterators[1]:
                            possible_col_names.append(f'{fill}{sep}{i}{sep}{j}')
                            possible_col_names.append(f'{i}{sep}{fill}{sep}{j}')
                            possible_col_names.append(f'{i}{sep}{j}{sep}{fill}')
            else:
                for sep in cls._COL_SEPARATORS:
                    for i in iterators[0]:
                        for j in iterators[1]:
                            for k in iterators[2]:
                                possible_col_names.append(f'{fill}{sep}{i}{sep}{j}{sep}{k}')
                                possible_col_names.append(f'{i}{sep}{fill}{sep}{j}{sep}{k}')
                                possible_col_names.append(f'{i}{sep}{j}{sep}{fill}{sep}{k}')
                                possible_col_names.append(f'{i}{sep}{j}{sep}{k}{sep}{fill}')
        return list(set(possible_col_names))

    @staticmethod
    def _check_logs(all_logs):
        log_error = False
        stem_height = 1
        for logs in all_logs:
            if logs[0] and logs[1]:
                continue
            elif not logs[0] and not logs[1]:
                log_error = True
            else:
                if not logs[0]:
                    stem_height += logs[1] + 1
                    logs[0] = stem_height
                else:
                    logs[1] = logs[0] - stem_height - 1
                    stem_height = logs[0]
        return log_error, all_logs

    @staticmethod
    def _error_float(value, required=True, non_negative=True):
        if not required and value in ['', ' ', None]:
            return False
        try:
            x = float(value)
            if non_negative and x < 0:
                return True
            else:
                return False
        except ValueError:
            return True

    @staticmethod
    def _error_int(value, required=True):
        if not required and value in ['', ' ', None]:
            return False
        else:
            try:
                x = int(value)
                if x < 0:
                    return True
                else:
                    return False
            except ValueError:
                return True

    @staticmethod
    def _error_string(value, required=True):
        if required and value in ['', ' ', None]:
            return True
        else:
            return False

    @staticmethod
    def _del_print_cols(cols):
        for col_name in cols:
            print(col_name)
            for sub in cols[col_name]:
                print(f'{sub}: {cols[col_name][sub]}')
            print()








if __name__ == '__main__':
    # file = 'C:/ZBEE490/Dev/other/timberpy/timberpy/del_scratch/csv/csv_stand_data_full_both.csv'
    # file = 'C:/ZBEE490/Dev/other/timberpy/timberpy/del_scratch/xlsx/stand_data_full_with_quick.xlsx'
    # file = 'C:/ZBEE490/Dev/other/timberpy/timberpy/del_scratch/xlsx/normal_to_try_multistand_no_hgt.xlsx'
    file = 'C:/ZBEE490/Dev/other/timberpy/timberpy/del_scratch/xls/stand_data_full_with_quick.xls'

    inv = Inventory(file)







