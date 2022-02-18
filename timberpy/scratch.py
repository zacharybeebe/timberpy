
if __name__ == '__main__':
    pass

    from random import randint, choice
    from openpyxl import Workbook

    WEST = (['DF'] * 50) + (['WH'] * 20) + (['RC'] * 18) + (['SS'] * 2) + (['RA'] * 5) + (['BM'] * 2) + (['CW'] * 3)
    EAST = (['PP'] * 50) + (['DF'] * 20) + (['RC'] * 5) + (['ES'] * 5) + (['WL'] * 10) + (['LP'] * 5) + (['BM'] * 2) + (['CW'] * 3)

    def get_dbh():
        return randint(200, 450) / 10

    def get_ht(dbh):
        hdr = randint(490, 730) / 10
        return (dbh / 12) * hdr

    def get_spp(side):
        return choice(side)

    stand = 'NRF_HQ'
    pf = 54.45
    name = 'del_scratch/fvs/NRF_HQ.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory'
    heads = ['STAND', 'PLOT', 'TREE', 'SPECIES', 'DBH', 'HEIGHT', 'PLOT FACTOR']
    for i, head in enumerate(heads, 1):
        ws.cell(1, i, head)

    row = 2
    for plot in range(1, 21):
        for tree in range(1, randint(5, 11)):
            dbh = get_dbh()
            temp = [stand, plot, tree, get_spp(WEST), dbh, get_ht(dbh), pf]
            for i, col in enumerate(temp, 1):
                ws.cell(row, i, col)
            row += 1

    wb.save(name)






